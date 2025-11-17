from django import forms
from django.contrib import admin, messages
from django.shortcuts import redirect, render
from django.urls import path
import csv
import io
from openpyxl import load_workbook

from .models import Payment, Promocode


@admin.register(Payment)
class PaymentAdmin(admin.ModelAdmin):
    fields = None


class PromocodeImportForm(forms.Form):
    file = forms.FileField(label='CSV/XLSX файл с промокодами')
    has_header = forms.BooleanField(
        label='Первая строка — заголовок', required=False, initial=True
    )


@admin.register(Promocode)
class PromocodeAdmin(admin.ModelAdmin):
    fields = ['denomination', 'promo', 'is_used']
    change_list_template = 'admin/pay/promocode/change_list.html'

    def get_urls(self):
        urls = super().get_urls()
        custom = [
            path('import/', self.admin_site.admin_view(self.import_view), name='pay_promocode_import'),
        ]
        return custom + urls

    def import_view(self, request):
        if request.method == 'POST':
            form = PromocodeImportForm(request.POST, request.FILES)
            if form.is_valid():
                file = form.cleaned_data['file']
                has_header = form.cleaned_data['has_header']
                try:
                    created, updated, skipped = self._import_file(file, has_header)
                except Exception as e:
                    messages.error(request, f'Ошибка импорта: {e}')
                    return redirect('admin:pay_promocode_changelist')
                if created or updated:
                    messages.success(
                        request,
                        f'Импорт завершён: создано {created}, обновлено {updated}, пропущено {skipped}.',
                    )
                else:
                    messages.warning(request, 'Импорт завершён, записей не найдено.')
                return redirect('admin:pay_promocode_changelist')
        else:
            form = PromocodeImportForm()
        context = dict(
            self.admin_site.each_context(request),
            opts=self.model._meta,
            form=form,
            title='Импорт промокодов из CSV/XLSX',
        )
        return render(request, 'admin/pay/promocode/import.html', context)

    def _import_file(self, uploaded, has_header: bool):
        name = uploaded.name.lower()
        created = updated = skipped = 0
        rows = []

        if name.endswith('.csv'):
            data = io.TextIOWrapper(uploaded.file, encoding='utf-8-sig')
            reader = csv.reader(data)
            rows = list(reader)
        elif name.endswith('.xlsx'):
            wb = load_workbook(io.BytesIO(uploaded.read()), read_only=True, data_only=True)
            ws = wb.active
            rows = [[cell if cell is not None else '' for cell in row] for row in ws.iter_rows(values_only=True)]
        else:
            raise ValueError('Поддерживаются только файлы .csv или .xlsx')

        if has_header and rows:
            rows = rows[1:]

        for idx, row in enumerate(rows, start=1):
            if not row:
                continue
            code = str(row[0]).strip()
            denom_raw = (row[1] if len(row) > 1 else '').__str__().strip()
            if not code:
                skipped += 1
                continue
            try:
                denomination = int(float(denom_raw)) if denom_raw else 0
            except ValueError:
                skipped += 1
                continue

            obj, is_created = Promocode.objects.update_or_create(
                promo=code,
                defaults={
                    'denomination': denomination,
                    # При импорте все промокоды считаются проданными,
                    # но ещё не использованными.
                    'is_sold': True,
                    'is_used': False,
                },
            )
            if is_created:
                created += 1
            else:
                updated += 1

        return created, updated, skipped
